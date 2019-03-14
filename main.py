import dataModule
import returnClasses as rc
import date_functions as dtf
import pandas as pd
import seaborn as sns
import matplotlib.pyplot as plt
import openpyxl

""""this module creates all output tables and graph and exports this to an excel file. Most arguments are shared by the
 functions below, they are set at right after the last function definition
    Shared Arguments:
        df(returnFrame): frame with all returns for calculations
        fund(string): name of the fund in df to be analyzed
        benchmark(string): name of the benchmark in df
        betas(string, list(string)): names of independent x variables for regression. Multiple names are input as a list
        yearList(list(int)): a list of integers resembling years, stats are calculated each year(int) period
        date(string): a month end date formatted as yyyy-mm-dd. This is the end date of the period that's being analyzed
"""


def get_returns(df, fund, benchmark, yearList, date):
    index = [str(i) + " year" for i in yearList]

    # create monthly returns dataframe
    output = {col: df.loc[:, col].month_return(date) for col in df.columns}
    df_returns_month = pd.DataFrame(output, index=['month'])

    # create ytd returns dataframe
    output = {col: df.loc[:, col].ytd_return(date) for col in df.columns}
    df_returns_ytd = pd.DataFrame(output, index=['ytd'])

    # create dataframe with  1,3, 5 year annualized returns for all funds in data
    output = {col: [df.loc[:, col].annualize(i, date)[0] for i in yearList] for col in df.columns}
    df_returns_ann = pd.DataFrame(output, index=index)

    # concatenate all Return dataframes, slice to fund and BM returns and add excess return
    dfR = pd.concat([df_returns_month, df_returns_ytd, df_returns_ann])
    dfR = dfR.loc[:, [fund, benchmark]]
    dfR['Excess'] = dfR[fund] - dfR[benchmark]
    return dfR


def get_riskstats(df, fund, benchmark, rf_rate, yearList, date):
    index = [str(i) + " year" for i in yearList]

    # create dataframe with 1,3,5 year TE
    output = {
        fund + '_TE': [df.loc[:, fund].tracking_error_annualized(df.loc[:, benchmark], i, date) for i in yearList]}
    df_te_ann = pd.DataFrame(output, index=index)

    # create dataframe with 1,3,5 year IR
    output = {fund + '_IR': [df.loc[:, fund].info_ratio_annualized(df.loc[:, benchmark], i, date) for i in yearList]}
    df_ir_ann = pd.DataFrame(output, index=index)

    # create dataframe with 1,3,5 year Sharpe ratio
    output = {fund + '_SR': [df.loc[:, fund].sharpe_annualized(df.loc[:, rf_rate], i, date) for i in yearList]}
    df_sr_ann = pd.DataFrame(output, index=index)

    # concatenate all Stats (risks) dataframes
    dfS = pd.concat([df_sr_ann, df_ir_ann, df_te_ann], axis=1)
    return dfS


def get_period_yearly_returns(df, fund, benchmark, yearList):
    """"create Period Yearly Returns (Pyr) in which daily returns are aggregated/compounded for each year"""

    start = df.index.max() - pd.offsets.YearBegin(n=max(yearList))
    dfPyr = df.loc[start:, [fund, benchmark]].period_returns('Y')
    dfPyr['Excess'] = dfPyr[fund] - dfPyr[benchmark]
    dfPyr.index = dfPyr.index.year  # convert date index to year
    dfPyr = dfPyr[::-1]  # reverse dateframe
    return dfPyr


def create_graph(df, fund, benchmark, yearList, date):
    # Compounded Return (Cr) graph fund vs benchmark ==========================
    start = df.index.max() - pd.offsets.YearBegin(n=max(yearList))

    # Create compounded returns dataframe for the period
    fund_compounded = df.loc[start:, fund].compounded_series(monthStartDate=start, monthEndDate=date)
    bm_compounded = df.loc[start:, benchmark].compounded_series(monthStartDate=start, monthEndDate=date)
    dfCr = pd.concat([fund_compounded, bm_compounded], axis=1).rename(columns={0: fund, 1: benchmark})

    # create graph
    sns.set(style='whitegrid')
    graph = sns.lineplot(data=dfCr, palette='tab10', linewidth=1)
    graph.lines[1].set_linestyle("-")

    yLabels = graph.get_yticks()
    yLabels = ['{:,.0%}'.format(y) for y in yLabels]
    graph.set_yticklabels(yLabels)
    graph.set(ylabel='Cumulative returns')

    graph.set_title('{} vs {}, {} year history'.format(fund, benchmark, max(yearList)))
    fig = graph.get_figure()
    fig.savefig('output.png')


def create_excel_file(fileName, sheetName):
    # ===== paste information to excel =====

    def format_table_values(ws, srow, table, custom_format):
        # this function formats the content of the tables that are being pasted in excel
        erow = table.shape[0] + 1 + srow
        if isinstance(table, pd.DataFrame):
            ecol = table.shape[1] + 1
        else:
            ecol = 2
        for row in ws.iter_rows(min_col=2, min_row=srow + 1, max_col=ecol, max_row=erow):
            for cell in row:
                cell.number_format = custom_format

    # configuring Excelwriter (openpyxl)
    try:
        wb = openpyxl.load_workbook(fileName)
    except:
        wb = openpyxl.Workbook()
        wb.save(fileName)

    writer = pd.ExcelWriter(fileName, engine='openpyxl')
    writer.book = wb  # add workbook to Pandas ExcelWriter.book attribute
    writer.sheets = dict(
        (ws.title, ws) for ws in wb.worksheets)  # add sheets dict to pandas ExcelWriter.sheets attribute

    # create a clean sheet by deleting prior
    if sheetName in [sheet.title for sheet in wb.worksheets]:
        wb.remove(wb[sheetName])

    wb.create_sheet(title=sheetName, index=0)
    ws = wb[sheetName]
    writer.sheets[sheetName] = ws  # add sheet dict to pandas ExcelWriter.sheets attribute

    # write dataframes to sheets

    srow = 1
    ws.cell(row=srow, column=1, value='Returns below with a minimal history of 1 year are annualized based on 251 days')
    dfR.to_excel(writer, sheet_name=sheetName, startrow=srow, startcol=0)
    format_table_values(ws, srow, dfR, '0.00%')
    srow += dfR.shape[0] + 4

    ws.cell(row=srow, column=1, value='Returns per yearly period')
    dfPyr.to_excel(writer, sheet_name=sheetName, startrow=srow, startcol=0)
    format_table_values(ws, srow, dfPyr, '0.00%')
    srow += dfPyr.shape[0] + 4

    ws.cell(row=srow, column=1, value='Sharpe Ratio, Information Ratio, Tracking Error')
    dfS.to_excel(writer, sheet_name=sheetName, startrow=srow, startcol=0)
    format_table_values(ws, srow, dfS, '0.00')
    srow += dfS.shape[0] + 4

    ws.cell(row=srow, column=1, value='Fama French 3 factor regression')
    regOutput.to_excel(writer, sheet_name=sheetName, startrow=srow, startcol=0)
    format_table_values(ws, srow, regOutput, '0.00')
    srow += regOutput.shape[0] + 2

    # convert r2 float to series to use .to_excel()
    r2srs = pd.Series(r2, index=['rsquared'], name="Multi factor")
    r2srs.to_excel(writer, sheet_name=sheetName, startrow=srow, startcol=0)
    format_table_values(ws, srow, r2srs, '0.00')
    srow += r2srs.shape[0] + 4

    ws.cell(row=srow, column=1, value='Benchmark regression')
    bmRegOutput.to_excel(writer, sheet_name=sheetName, startrow=srow, startcol=0)
    format_table_values(ws, srow, bmRegOutput, '0.00')
    srow += bmRegOutput.shape[0] + 2

    # convert r2 float to series to use .to_excel()
    bmR2srs = pd.Series(bmR2, index=['rsquared'], name="Benchmark")
    bmR2srs.to_excel(writer, sheet_name=sheetName, startrow=srow, startcol=0)
    format_table_values(ws, srow, bmR2srs, '0.00')
    srow += bmR2srs.shape[0] + 3

    ws.cell(row=srow, column=1, value='Information for period ending on ' + str(date4xl) +
                                       '. Regressions are based on a 5 year history.')

    # paste graph to excel file
    img = openpyxl.drawing.image.Image('output.png')
    ws.add_image(img, 'I1')

    # save writer/excel file
    writer.save()


if __name__ == '__main__':
    # declare string variables according to column names in dataframe
    fund = 'Fund'  # string
    benchmark = 'BM'  # string
    betas = ['Mkt-RF', 'SMB', 'HML']  # string or list, used to determine columns for X variables in regression
    rf_rate = 'RF'  # string
    y = 'y'  # string used to determine column used for y in regression

    # date determines the end date of the period over which the calculations are performed
    date = 'lastMonthEnd'

    # each int in yearList determines the period timespan going back from date
    yearList = [1, 3, 5]

    # variables to determine sheet and file output names
    sheetName = 'factsheet'
    fileName = 'factsheet.xlsx'

    # ===== no more declarations code executes below ======

    # get data
    df = dataModule.get_data()
    date4xl = dtf.get_month_end(df, date=date).date()

    # add column for y variable in regression
    df.loc[:, y] = df.loc[:, fund] - df.loc[:, rf_rate]

    # Create output dataframes
    dfS = get_riskstats(df=df, fund=fund, benchmark=benchmark, rf_rate=rf_rate, yearList=yearList, date=date)
    dfR = get_returns(df=df, fund=fund, benchmark=benchmark, yearList=yearList, date=date)
    dfPyr = get_period_yearly_returns(df=df, fund=fund, benchmark=benchmark, yearList=yearList)

    # Create graph
    create_graph(df=df, fund=fund, benchmark=benchmark, yearList=yearList, date=date)

    # perform multifactor regression regression over the longest year period in yearList, y = fund returns - rf
    regOutput, r2 = df.regression(y, betas, max(yearList), date)

    # perform benchmark regression over the longest year period in yearList, y = fund returns
    bmRegOutput, bmR2 = df.regression(fund, benchmark, max(yearList), date)

    # Create excel file
    create_excel_file(fileName=fileName, sheetName=sheetName)
